# Auto Verrification Tool - build with love :3
# -----------------------------------------------------------------------------------------------
# Simple command-line tool for automate the process of creating QT overlay verification document.
# -----------------------------------------------------------------------------------------------
# Author: David Wang
# Copyright 2017 Quanta Inc

import argparse
import openpyxl
import csv
import string
import getpass
import re
import copy
from datetime import date, datetime
from openpyxl.worksheet import *
from openpyxl.styles import Font, Border, Side, Color, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_from_string
from openpyxl.cell import Cell

# Set up arguments for this program
parser = argparse.ArgumentParser(description="Auto Verification program!", formatter_class=argparse.RawTextHelpFormatter)
parser.add_argument("-r", "--reviser", default=string.capwords(getpass.getuser()), help="who release this verification document")
parser.add_argument("-w", "--reviewer", default="Doris", help="the reviewer name, default name is Doris")
parser.add_argument("-s", "--source", required=True, help="path of source .xlsx verification document")
parser.add_argument("-l", "--csvlog", required=True, help="path of .csv file in CSVLOG folder")
parser.add_argument("-m", "--modem", required=True, help="path of .txt file in MODEM folder")
parser.add_argument("-c", "--csv", required=True, help="path of .csv file in CSV folder")
parser.add_argument("-v", "--ver", required=True, help="version number of the new test plan")
parser.add_argument("-d", "--date", required=True, help="overlay verified date, format: 20170509")

# Combine all arguments into a list called args
args = parser.parse_args()
reviser_name = args.reviser
reviewer_name = args.reviewer
open_workbook = args.source # source Excel file path
open_csvlog = args.csvlog # CSVLOG file path
open_modem = args.modem # MODEM log file path
open_csv = args.csv # CSV file path
testplan_ver = args.ver
ovl_verify_date = args.date

# Cell background color options
rgb_black = [0,0,0] # Black color
black_color_string = "".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb_black])

rgb_gray = [190,190,190] # Gray color
gray_color_string = "".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb_gray])

rgb_lightgray = [211,211,211] # Light Gray
lightgray_color_string = "".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb_lightgray])

rgb_white = [255,255,255] # White color
white_color_string = "".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb_white])

rgb_cyan = [176,196,222] # Cyan Blue
cyan_color_string = "".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb_cyan])

rgb_navy = [25,25,112] # Navy Blue
navy_color_string = "".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb_navy])

def set_border(ws, cell_range, thin_border=True, color=black_color_string):
    rows = ws[cell_range]
    side = Side(border_style='medium', color=black_color_string) # Black color border by default

    rows = list(rows)
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            if thin_border:
                border = Border(left=Side(style='thin', color=color),
                                right=Side(style='thin', color=color),
                                top=Side(style='thin', color=color),
                                bottom=Side(style='thin', color=color))
            else:
                border = Border(left=Side(),
                                right=Side(),
                                top=Side(),
                                bottom=Side())
                
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # Set thin border for each cell
            cell.border = border

            # set medium border for the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border
                
def set_cells_color(cell_range, select_color):
    for cIndex1, cells in enumerate(worksheet[cell_range]):
        for cIndex2, cell in enumerate(cells):
            cell.fill = PatternFill(fill_type="solid", start_color='FF' + select_color, end_color='FF' + select_color)

def set_font_style(cell_range, font_bold=True, font_italic=False, font_color="000000", font_type="Times New Roman"):
    for cIndex1, cells in enumerate(worksheet[cell_range]):
        for cIndex2, cell in enumerate(cells):
            cell.font = Font(color = font_color)
            cell.font = cell.font.copy(name=font_type, bold=font_bold, italic=font_italic)
            
def is_number(s):
    try:
        float(s) # for int, long and float
    except ValueError:
        try:
            complex(s) # for complex
        except ValueError:
            return False
    return True

def clear_extra_cells(cell_range):
    border = Border(left=Side(),right=Side(), top=Side(), bottom=Side())
    for cIndex1, cells in enumerate(worksheet[cell_range]):
        for cIndex2, cell in enumerate(cells):
            cell.fill = PatternFill()
            cell.border = border

def set_column_width(start_cell_letter, end_cell_letter, column_width, larger_one):
    start_index = column_index_from_string(start_cell_letter)
    end_index = column_index_from_string(end_cell_letter)
    numbers_of_col = (end_index - start_index) + 1
    if numbers_of_col < larger_one:
           column_width = int((column_width * larger_one) / numbers_of_col)
    for i in range(start_index, end_index + 1):
        worksheet.column_dimensions[get_column_letter(i)].width = column_width
    worksheet.column_dimensions[get_column_letter(end_index + 1)].width = 5

def insert_rows(self, row_idx, cnt, above=False, copy_style=True, copy_merged_columns=True, fill_formulae=True):
    CELL_RE = re.compile("(?P<col>\$?[A-Z]+)(?P<row>\$?\d+)")
    row_idx = row_idx - 1 if above else row_idx
    def replace(m):
        row = m.group('row')
        prefix = "$" if row.find("$") != -1 else ""
        row = int(row.replace("$", ""))
        row += cnt if row > row_idx else 0
        return m.group('col') + prefix + str(row)

    # First, we shift all cells down cnt rows...
    old_cells = set()
    old_fas = set()
    new_cells = dict()
    new_fas = dict()
    for c in self._cells.values():

        old_coor = c.coordinate

        # Shift all references to anything below row_idx
        if c.data_type == Cell.TYPE_FORMULA:
            c.value = CELL_RE.sub(
                replace,
                c.value
            )
            # Here, we need to properly update the formula references to reflect new row indices
            if old_coor in self.formula_attributes and 'ref' in self.formula_attributes[old_coor]:
                self.formula_attributes[old_coor]['ref'] = CELL_RE.sub(
                    replace,
                    self.formula_attributes[old_coor]['ref']
                )

        # Do the magic to set up our actual shift
        if c.row > row_idx:
            old_coor = c.coordinate
            old_cells.add((c.row, c.col_idx))
            c.row += cnt
            new_cells[(c.row, c.col_idx)] = c
            if old_coor in self.formula_attributes:
                old_fas.add(old_coor)
                fa = self.formula_attributes[old_coor].copy()
                new_fas[c.coordinate] = fa

    for coor in old_cells:
        del self._cells[coor]
    self._cells.update(new_cells)

    for fa in old_fas:
        del self.formula_attributes[fa]
    self.formula_attributes.update(new_fas)

    # Next, we need to shift all the Row Dimensions below our new rows down by cnt...
    # CHANGED: for row in range(len(self.row_dimensions) - 1 + cnt, row_idx + cnt, -1):
    for row in range(list(self.row_dimensions)[-1] + cnt, row_idx + cnt, -1):
        new_rd = copy.copy(self.row_dimensions[row - cnt])
        new_rd.index = row
        self.row_dimensions[row] = new_rd
        del self.row_dimensions[row - cnt]

    # Now, create our new rows, with all the pretty cells
    # CHANGED: row_idx += 1
    new_row_idx = row_idx + 1
    for row in range(new_row_idx, new_row_idx + cnt):
        # Create a Row Dimension for our new row
        new_rd = copy.copy(self.row_dimensions[row-1])
        new_rd.index = row
        self.row_dimensions[row] = new_rd

        # Set row height of the new row
        rowHeights = [self.row_dimensions[i+1].height for i in range(self.max_row)]
        rowHeights = [15 if rh is None else rh for rh in rowHeights]
        self.row_dimensions[row].height = max(rowHeights)

        # CHANGED: for col in range(1,self.max_column):
        for col in range(self.max_column):
            # CHANGED: col = get_column_letter(col)
            col = get_column_letter(col+1)
            cell = self['%s%d' % (col, row)]
            cell.value = None
            source = self['%s%d' % (col, row+1)]
            if copy_style:
                cell.number_format = source.number_format
                cell.font = source.font.copy()
                cell.alignment = source.alignment.copy()
                cell.border = source.border.copy()
                cell.fill = source.fill.copy()
            if fill_formulae and source.data_type == Cell.TYPE_FORMULA:
                s_coor = source.coordinate
                if s_coor in self.formula_attributes and 'ref' not in self.formula_attributes[s_coor]:
                    fa = self.formula_attributes[s_coor].copy()
                    self.formula_attributes[cell.coordinate] = fa
                #print("Copying formula from cell %s%d to %s%d"%(col,row-1,col,row))
                cell.value = re.sub(
                    "(\$?[A-Z]{1,3}\$?)%d" % (row-1),
                    lambda m: m.group(1) + str(row),
                    source.value
                )
                cell.data_type = Cell.TYPE_FORMULA

    # Check for Merged Cell Ranges that need to be expanded to contain new cells
    for cr_idx, cr in enumerate(self.merged_cell_ranges):
        self.merged_cell_ranges[cr_idx] = CELL_RE.sub(
            replace,
            cr
        )

    # Merge columns of the new rows in the same way row above does
    if copy_merged_columns:
        for cr in self.merged_cell_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(cr)
            if max_row == min_row == row_idx:
                for row in range(new_row_idx, new_row_idx + cnt):
                    newCellRange = get_column_letter(min_col) + str(row) + ":" + get_column_letter(max_col) + str(row)
                    self.merge_cells(newCellRange)
                    
Worksheet.insert_rows = insert_rows


# Load the source .xlsx template file
workbook = openpyxl.load_workbook(open_workbook)

print("\nOpening source workbook '{}'\n".format(open_workbook.split('/')[-1]))


# 0. Fetch some of the required data

with open(open_csvlog, newline='', encoding='utf_8') as f:
    reader = csv.reader(f)
    for r, row in enumerate(reader):
        for c, col in enumerate(row):
            if r is 0 and c is 0: # Get station name
                station_name = col
            #elif r is 0 and c is 1: # Get hashtag
                #unsplit_ovl_name = col.strip('SW_Version:').split('_')
                #ovl_version_name = ["{}_{}_{}".format(unsplit_ovl_name[unsplit_ovl_name.index(i) - 1],
                                                      #unsplit_ovl_name[unsplit_ovl_name.index(i)],
                                                      #unsplit_ovl_name[unsplit_ovl_name.index(i) + 1])
                                    #for i in unsplit_ovl_name if 'JH' in i].pop()
                #csvlog_date_str = ovl_version_name[ovl_version_name.index('ver')-8:ovl_version_name.index('ver')]
                #unsplit_ovl_name = col.split(':')
                #hashtag = unsplit_ovl_name[1].strip('V')[:7]
            elif r is 0 and c is 2:
                serial_num = col.strip('Serial Number:')
            elif col == "DIAGS_VERSION":
                diags_version = row[c + 3] # Get the DIAGS_VERSION value
            elif "total test time" in col.lower():
                total_test_time = row[c + 1] # Get the total test time value

# Create new version name for verification document
worksheet = workbook.get_sheet_by_name('CSV log comparison')
version_name_comp = worksheet['D2'].value.split(':')[1].strip(' ').split('_')
qtm_version_num = str(int(version_name_comp[2]) + 1).zfill(3) # or use '%0*d' % (3, 10)
new_version_name = version_name_comp[0].replace(version_name_comp[0][version_name_comp[0].index('ver')-8:version_name_comp[0].index('ver')], ovl_verify_date).replace(version_name_comp[0][version_name_comp[0].index('ver')+3:], testplan_ver)
ovl_version_name = new_version_name + '_' + version_name_comp[1] + '_' + qtm_version_num

# Get date of today in a particular format
release_date = datetime.strftime(date.today(), "%Y/%m/%d")
print('---------------------------------------')
print('Station Name:', station_name)
print('Version Name:', ovl_version_name)
print('Verify Date:', datetime.strptime(ovl_verify_date, '%Y%m%d').strftime('%Y/%m/%d'))
print('Test Sample SN:', serial_num)
print('Reviser:', reviser_name)
print('Reviewer:', reviewer_name)
print('Release Date', release_date)
print('---------------------------------------\n')

# 1. Creating data in 'Version' worksheet
worksheet = workbook.get_sheet_by_name('Version ')
            
print("[Step 1] Creating data in 'Version' worksheet...\n")

worksheet['B3'] = ovl_version_name # Version
worksheet['C5'] = ovl_version_name # Overlay Version
worksheet['C3'] = release_date # Release Date (date of today by default)
worksheet['D3'] = station_name # Station Name
worksheet['E3'] = reviser_name # Reviser
worksheet['C6'] = datetime.strptime(ovl_verify_date, '%Y%m%d').strftime('%Y/%m/%d') # Verify Date
worksheet['C8'] = reviewer_name # Reviewer ('Doris' by default)
worksheet['C9'] = serial_num # Test Sample SN

# Apply border to selected range of cells
set_border(worksheet, "B2:E3")
set_border(worksheet, "B5:E9")
set_border(worksheet, "B11:D15")

print("Complete creating data in 'Version' worksheet\n")


# 2.Creating data in 'Program Verification' sheet
print("[Step 2] Creating data in 'Program Verification' worksheet...\n")

worksheet = workbook.get_sheet_by_name('Program Verification')
worksheet.insert_rows(3, 1, above=True, copy_style=True, copy_merged_columns=True, fill_formulae=True)
# Fill data into each cell
worksheet['B3'] = release_date # Date
worksheet['C3'] = station_name # Test station
worksheet['F3'] = "N/A" # Fail symptom
worksheet['G3'] = "Version:\n{}\nDiag Ver: {}".format(ovl_version_name, diags_version) # Program Version 
worksheet['I3'] = float(total_test_time) # Test time

print("Complete creating data in 'Program Verification' worksheet\n")


# 3.Creating data in 'CSV log comparison' sheet
print("[Step 3] Creating data in 'CSV log comparison' worksheet...\n")

worksheet = workbook.get_sheet_by_name('CSV log comparison')

csvfile = open_csvlog

print("Reading csvlog data from {}...\n".format(open_csvlog.split('/')[-1]))

# Clear data in right-hand side cell section
cell_range = 'K{}:P{}'.format(worksheet.min_row, worksheet.max_row)
rows = worksheet[cell_range]
rows = list(rows)
border = Border(
    left=Side(),
    right=Side(),
    top=Side(),
    bottom=Side()
)
lastCol = 5 # Use the default value, if right-side cell section is blank
for pos1, cells in enumerate(rows):
    for pos2, cell in enumerate(cells):
        if "total test time" in str(cell.value).lower():
            lastCol = pos1 + 1
            # Deleting data of cell range
            for row in worksheet['K4:P{}'.format(lastCol)]:
                for cell in row:
                    cell.value = None
                    cell.border = border
                    
# Set gray color for right-hand cell section
#set_cells_color("K4:P{}".format(lastCol), gray_color_string)

# Get total row counts of source CSVLOG file
with open(csvfile, newline='', encoding='utf_8') as f:
    reader = csv.reader(f)
    row_count = len(list(reader))
    
# Set gray color to background section
margin_bottom_rows = 2 # Can customize the margin bottom row value
cell_range = "A16:Q{}".format(max(lastCol, row_count + 3) + margin_bottom_rows)
set_cells_color(cell_range, gray_color_string)

# Copy left CSVLOG to right-hand side cell sections
cell_range = 'D{}:I{}'.format(worksheet.min_row, worksheet.max_row)
rows = worksheet[cell_range]
rows = list(rows)
for pos1, cells in enumerate(rows):
    for pos2, cell in enumerate(cells):
        if "total test time" in str(cell.value).lower():
            lastCol = pos1 + 1

# Save values from cell range A to a list
cell_range = 'D4:I{}'.format(lastCol)
rows = worksheet[cell_range]
rows = list(rows)
translated = []
border = Border(left=Side(),
                right=Side(),
                top=Side(),
                bottom=Side()
                )
for pos1, cells in enumerate(rows):
    for pos2, cell in enumerate(cells):
        translated.append(cell.value)
        # Clear data and border setting in left-hand side cells
        cell.value = None
        cell.border = border

# Set gray color for left-hand cell section
set_cells_color(cell_range, gray_color_string)

# Paste values from list to cell range B
cell_range = 'K4:P{}'.format(lastCol)
rows = worksheet[cell_range]
rows = list(rows)
for pos1, cells in enumerate(rows):
    for pos2, cell in enumerate(cells):
        cell.value = translated.pop(0)
        # Fill white color background to each cell
        cell.fill = PatternFill(fill_type="solid", start_color='FF' + white_color_string, end_color='FF' + white_color_string)

# Copy header value from ws['D2'] to ws['K2']
worksheet['K2'].value = worksheet['D2'].value

# Apply border style to right-hand cell section
set_border(worksheet, "K2:P2")
set_border(worksheet, "K3:P3")
set_border(worksheet, "K4:P{}".format(lastCol - 1), True, lightgray_color_string)
set_border(worksheet, "K{}:P{}".format(lastCol, lastCol), False, lightgray_color_string)
set_cells_color("K{}:P{}".format(lastCol, lastCol), cyan_color_string)

# Copy data from source CSVLOG file to left-hand cell section
with open(csvfile, newline='', encoding='utf_8') as f:
    reader = csv.reader(f)
    for r, row in enumerate(reader):
        for c, col in enumerate(row):
            this_cell = worksheet.cell(row = r+4, column = c+4)
            if col != 'nan' and is_number(col): # Check if data is number
                if '.' in col: # Convert string to float, if data is float number
                    this_cell.value = float(col)
                else:
                    this_cell.value = int(col)
            else:
                this_cell.value = col
            # Set all cells in left-side section to have same font type
            this_cell.font = Font(color = "000000")
            this_cell.font = this_cell.font.copy(name='Times New Roman', bold=False, italic=False)

# Apply white background color and set border to left-hand cell section
set_cells_color("D4:I{}".format(3 + row_count - 1), white_color_string)
set_border(worksheet, "D4:I{}".format(3 + row_count - 1), True, lightgray_color_string)

# Apply blue background color and border to left-hand bottom cells (Total Test Time)
set_border(worksheet, "D{}:I{}".format(3 + row_count, 3 + row_count), False, lightgray_color_string)
set_cells_color("D{}:I{}".format(3 + row_count, 3 + row_count), cyan_color_string)
set_font_style("D{}:I{}".format(3 + row_count, 3 + row_count))

# Set data and border for 2 header cells in left-hand cell section
stationVersion= '{} VERSION: {}'.format(station_name, ovl_version_name)
worksheet['D2'] = stationVersion
worksheet['D3'] = 'CSV LOG'
set_border(worksheet, "D2:I2")
set_border(worksheet, "D3:I3")

# If max_row (gray background section) larger then data cell section, delete extra gray rows
if worksheet.max_row > max(lastCol, row_count + 3) + 1:
    extraCellstartIndex = max(lastCol, row_count + 3) + 2
    extraCellEndIndex = worksheet.max_row
    grayCellMargin = worksheet.max_column
    clear_extra_cells("A{}:{}{}".format(extraCellstartIndex, get_column_letter(grayCellMargin), extraCellEndIndex))
       
print("Complete creating data in 'CSV log comparison' worksheet\n")


# 4. Creating data in 'UART Log Check' sheet
print("[Step 4] Creating data in 'UART Log Check' worksheet...\n")

worksheet = workbook.get_sheet_by_name('UART Log Check')

# Clear the data in right-side cell section
search_range = "A3:{}3".format(get_column_letter(worksheet.max_column - 1))
rows = worksheet[search_range]
rows = list(rows)
for pos1, cells in enumerate(rows):
    for pos2, cell in enumerate(cells):
        if str(cell.value).lower() == "uart log":
            right_cell_start_index = pos2 + 1

rightSectionStartIndex = get_column_letter(right_cell_start_index) # Letter of right-side cell start idex
rightSectionEndIndex = get_column_letter(worksheet.max_column-1) # Letter of right-side cell end idex
leftSectionEndIndex = get_column_letter(right_cell_start_index - 2) # Letter of left-side cell end idex

rightSectionWidth = worksheet.max_column - right_cell_start_index
leftSectionWidth = right_cell_start_index - 5

cell_range = '{}4:{}{}'.format(rightSectionStartIndex, rightSectionEndIndex, worksheet.max_row)
rows = worksheet[cell_range]
rows = list(rows)
border = Border(
    left=Side(),
    right=Side(),
    top=Side(),
    bottom=Side()
)
lastCol = 5 # Use the default value, if right-side cell section is blank
for pos1, cells in enumerate(rows):
    for pos2, cell in enumerate(cells):
        if cell.value is not None:
            lastCol = pos1 + 5 # Index of last row of right-hand cell section
rightSectionRange = '{}2:{}{}'.format(rightSectionStartIndex, rightSectionEndIndex, worksheet.max_row)
for row in worksheet[rightSectionRange]:
    for cell in row:
        cell.value = None
        cell.border = border
        cell.fill = PatternFill()

# If width of left-side section is larger the right-side section, adjust width of right-side section
right_cell_add_cols = 0
if leftSectionWidth > rightSectionWidth:
    right_cell_add_cols = leftSectionWidth - rightSectionWidth

f = open(open_modem, 'r+')
data = f.readlines() # read all lines at once
print("Reading Modem log data from '{}'\n".format(open_modem.split('/')[-1]))

# Counting max row width of data in the source modem log
modem_max_width = 0
for i in range(len(data)): # i = row index, j = column index, row = cell value
    row = data[i].split('\t')
    for j in range(len(row)):
        if len(row) > modem_max_width:
            modem_max_width = len(row)

# If max width of data in modem log is wider than left-side section, adjust the width of gray background section
left_cell_add_cols = 0
if modem_max_width > leftSectionWidth:
    left_cell_add_cols = modem_max_width - leftSectionWidth

# Set gray color for adjusted extra cell section
extra_gray_range = "{}1:{}{}".format(get_column_letter(worksheet.max_column + 1), get_column_letter(worksheet.max_column + right_cell_add_cols + left_cell_add_cols), worksheet.max_row)
set_cells_color(extra_gray_range, gray_color_string)

# Set gray color for right-side cell section
set_cells_color(rightSectionRange, gray_color_string)

# Copy left Modem log to right-side cell section
cell_range = 'D4:{}{}'.format(leftSectionEndIndex, worksheet.max_row)
rows = worksheet[cell_range]
rows = list(rows)
for pos1, cells in enumerate(rows):
    for pos2, cell in enumerate(cells):
        if cell.value is not None:
            lastCol = pos1 + 5

# Store data from cell section A to a list
cell_range = 'D4:{}{}'.format(leftSectionEndIndex, lastCol)
rows = worksheet[cell_range]
rows = list(rows)
translated = []
for pos1, cells in enumerate(rows):
    for pos2, cell in enumerate(cells):
        if cell.value is None:
            cell.value = ''
        translated.append(cell.value)
        # Clear data and border setting in left-hand side cells
        cell.value = None
        cell.border = border

# Set gray color for left-hand cell section
set_cells_color(cell_range, gray_color_string)

# Set gray color to background section
# If max_row of gray background is larger then data cell section, delete extra gray rows
if worksheet.max_row > max(lastCol, len(data) + 4) + 1:
    extraCellstartIndex = max(lastCol, len(data) + 4) + 2
    extraCellEndIndex = worksheet.max_row
    grayCellMargin = worksheet.max_column
    clear_extra_cells("A{}:{}{}".format(extraCellstartIndex - 1, get_column_letter(grayCellMargin), extraCellEndIndex))
    set_cells_color("A{}:{}{}".format(extraCellstartIndex - 1, get_column_letter(grayCellMargin), extraCellstartIndex - 1), gray_color_string)
else:
    # If max_row of gray background is shorter then data cell section, fill in gray background
    margin_bottom_rows = 1 # Can customize the margin bottom row value
    cell_range = "A{}:{}{}".format(worksheet.max_row, get_column_letter(worksheet.max_column), max(lastCol, len(data) + 4) + margin_bottom_rows)
    set_cells_color(cell_range, gray_color_string)

# Adjust bandwidth of the two cell sections
rightSectionEndIndex = get_column_letter(worksheet.max_column - 1)
if left_cell_add_cols is not 0:
    # Adjust the start index of right-side cell section
    rightSectionStartIndex = get_column_letter(right_cell_start_index + left_cell_add_cols)
    # Adjust the end index of left-side cell section
    leftSectionEndIndex = get_column_letter(right_cell_start_index - 2 + left_cell_add_cols)

# Paste data from list to right-side cell section
cell_range = '{}4:{}{}'.format(rightSectionStartIndex, rightSectionEndIndex, lastCol)
rows = worksheet[cell_range]
rows = list(rows)
for pos1, cells in enumerate(rows):
    for pos2, cell in enumerate(cells):
        cell.value = translated.pop(0)
        # Fill white color background to each cell
        cell.fill = PatternFill(fill_type="solid", start_color='FF' + white_color_string, end_color='FF' + white_color_string)

# Copy header value from left-side to right-side section, then set border style and cell color
header_cell1 = worksheet['{}2'.format(rightSectionStartIndex)]
header_cell2 = worksheet['{}3'.format(rightSectionStartIndex)]
header_cell1.value = worksheet['D2'].value
header_cell2.value = worksheet['D3'].value
right_section_header_range1 = '{}2:{}2'.format(rightSectionStartIndex, rightSectionEndIndex)
right_section_header_range2 = '{}3:{}3'.format(rightSectionStartIndex, rightSectionEndIndex)
set_border(worksheet, right_section_header_range1, False)
set_border(worksheet, right_section_header_range2, False)
set_cells_color(right_section_header_range1, navy_color_string)
set_cells_color(right_section_header_range2, cyan_color_string)

# Set font color to white and bold style for two header columns
header_cell1.font = Font(color = "FFFFFF")
header_cell1.font = header_cell1.font.copy(name='Times New Roman', bold=True, italic=False)
header_cell2.font = Font(color = "000000")
header_cell2.font = header_cell2.font.copy(name='Times New Roman', bold=True, italic=False, size=14)

# Apply border style to right-hand cell section
set_border(worksheet, cell_range, False)

# Copy data from source Modem log file into left-hand cell section
for i in range(len(data)): # i = row index, j = column index, row = cell value
    # This will return a line of string data, may need to convert to other formats depending on use case
    row = data[i].split('\t')
    for j in range(len(row)):
        # Check and remove illegal character (ASCII)
        filtered_string = "".join(filter(lambda x: x in string.printable, row[j]))
        # Convert cell value that contains multiple "=" to plan text string
        if "==" in filtered_string:
            filtered_string = " " + filtered_string
        # Write value to cell
        worksheet.cell(row = i+4, column = j+4).value = filtered_string

# Apply white background color and set border to left-hand cell section
cell_range = 'D4:{}{}'.format(leftSectionEndIndex, len(data) + 4)
set_cells_color(cell_range, white_color_string)
set_border(worksheet, cell_range, False)

# Set header value for left-side section, and set border style and cell color
worksheet['D2'].value = stationVersion
left_section_header_range1 = 'D2:{}2'.format(leftSectionEndIndex)
left_section_header_range2 = 'D3:{}3'.format(leftSectionEndIndex)
set_border(worksheet, left_section_header_range1, False)
set_border(worksheet, left_section_header_range2, False)
set_cells_color(left_section_header_range1, navy_color_string)
set_cells_color(left_section_header_range2, cyan_color_string)
    
# Set each column to same width and set two cell sections have same wide size
rightSectionWidth = worksheet.max_column - right_cell_start_index
leftSectionWidth = right_cell_start_index - 5
col_width = 50 # The default cell width can be adjusted
larger_one = max(rightSectionWidth, leftSectionWidth)
set_column_width('D', leftSectionEndIndex, col_width, larger_one)
set_column_width(rightSectionStartIndex, rightSectionEndIndex, col_width, larger_one)
       
print("Complete creating data in 'UART Log Check' worksheet\n")


# 5. Creating data in 'CSV file' sheet
print("[Step 5] Creating data in 'CSV file' worksheet...\n")

# Delete current worksheet and create a new one
workbook.remove_sheet(workbook.get_sheet_by_name('CSV file'))
workbook.create_sheet('CSV file')
worksheet = workbook.get_sheet_by_name('CSV file')

csvfile = open_csv

print("Copying data from '{}'...\n".format(open_csv.split('/')[-1]))

# Then write data into this worksheet from csv file
with open(csvfile, newline='', encoding='utf_8') as f:
    reader = csv.reader(f)
    for r, row in enumerate(reader):
        for c, col in enumerate(row):
            worksheet.cell(row = r+1, column = c+1).value = col

print("Complete creating data in 'CSV file' worksheet\n")


# 6. Save changes to the created new verification file
target_filename = "QTVerification_{}_{}.xlsx".format(station_name, ovl_version_name)
workbook.save(target_filename)

print("Verification document '{}' is successfully created!\n".format(target_filename))
